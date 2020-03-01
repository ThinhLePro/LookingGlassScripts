from splinter import Browser
from selenium import webdriver
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import openpyxl,time,threading,datetime,os,ssl,smtplib

def GetLgHeNet(LstIP,browser,DctResult):
    for Ip in LstIP:
        StrResult  = 'None'
        print('GetLgHeNet : %s .Inprogress : %d/%d'%(Ip,LstIP.index(Ip)+1,len(LstIP)))
        try:
            time.sleep(5)
            browser.visit('https://lg.he.net/')
            time.sleep(5)
            browser.click_link_by_id('command_ping')
            browser.find_by_id('ip').fill(Ip.strip())
            browser.find_by_id('raw').click()
            browser.find_by_value('Probe').click()
            time.sleep(5)
            while True:
                if browser.is_text_present('You have too many active queries'): 
                    DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
                    receiver_email = 'thinhlv@vng.com.vn'
                    subject = '[Ping monitor tool] Send error message'
                    message = '[Function : Main][Time collect: %s][Error : You have too many active queries]'%(DateTimeCurrent)
                    SendEmailText(receiver_email,message,subject)
                    time.sleep(60)
                    browser.find_by_value('Probe').click()
                    time.sleep(5)
                else: 
                    break
            StrResult = browser.find_by_id('lg_return').text
        except Exception as error : 
            print('GetLgHeNet : %s\n%s'%(Ip,error))
            DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
            receiver_email = 'thinhlv@vng.com.vn'
            subject = '[Ping monitor tool] Send error message'
            message = '[Function : Main][Time collect: %s][Error : %s]'%(DateTimeCurrent,error)
            SendEmailText(receiver_email,message,subject)
            time.sleep(60)
        
        if Ip in DctResult: DctResult[Ip]['LgHeNet'] = StrResult
        else: DctResult[Ip] = {'LgHeNet':StrResult}



    
def GetCenturyLink(LstIP,browser,DctResult):
    for Ip in LstIP:
        StrResult  = 'None'
        print('GetCenturyLink : %s .Inprogress : %d/%d'%(Ip,LstIP.index(Ip)+1,len(LstIP)))
        try:
            time.sleep(5)
            browser.visit('https://lookingglass.centurylink.com/')
            time.sleep(5)
            browser.find_by_value('Singapore').click()
            browser.find_by_xpath('//div[@class="col-12 col-sm-9 col-md-6 col-lg-6 col-xl-6"]/input').first.fill(Ip)
            browser.find_by_xpath('//div[@class="col-7 col-sm-7 col-md-6 col-lg-6 col-xl-6"]/input').first.fill('32')
            browser.find_by_xpath('//div[@class="col-7 col-sm-7 col-md-6 col-lg-6 col-xl-6"]/select').last.select('5')
            browser.find_by_xpath('//button[@class="btn  btn-primary btn-sm"]').first.click()
            time.sleep(5)
            StrResult = browser.find_by_xpath('//div[@class="container-fluid"]/div[3]/div[2]').text
        except Exception as error : 
            print('GetCenturyLink : %s\n%s'%(Ip,error))
            DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
            receiver_email = 'thinhlv@vng.com.vn'
            subject = '[Ping monitor tool] Send error message'
            message = '[Function : Main][Time collect: %s][Error : %s]'%(DateTimeCurrent,error)
            SendEmailText(receiver_email,message,subject)
            time.sleep(60)

        if Ip in DctResult: DctResult[Ip]['CenturyLink'] = StrResult
        else: DctResult[Ip] = {'CenturyLink':StrResult}

def GetPCCW(LstIP,browser,DctResult):
    for Ip in LstIP:
        StrResult  = 'None'
        print('GetPCCW : %s .Inprogress : %d/%d'%(Ip,LstIP.index(Ip)+1,len(LstIP)))
        try:
            time.sleep(5)
            browser.visit('https://lookingglass.pccwglobal.com/')
            time.sleep(5)
            browser.find_by_xpath('//div[@id="srcContainer"]/select/option[@value="sin01"]').click()
            browser.find_by_xpath('//div[@id="rProtocolContainer"]/select/option[@value="ipv4"]').click()
            browser.find_by_xpath('//div[@id="rProfileContainer"]/select/option[@value="standard"]').click()
            browser.find_by_id('offNet').first.click()
            browser.find_by_id('newOffNet').first.fill(Ip)
            browser.find_by_id('addOff').first.click()
            #browser.find_by_xpath('//div[@id="serviceContainer"]/select/option[@value="ping"]').click()
            #browser.find_by_xpath('//div[@id="packetCountContainer"]/select/option[@value="100"]').click()
            browser.find_by_id('submit').first.click()
            while True:
                if browser.is_text_present('Query Complete'): break
                elif browser.is_text_present('Request has timed out'): 
                    DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
                    receiver_email = 'thinhlv@vng.com.vn'
                    subject = '[Ping monitor tool] Send error message'
                    message = '[Function : Main][Time collect: %s][Error : Request has timeout]'%(DateTimeCurrent)
                    SendEmailText(receiver_email,message,subject)
                    browser.find_by_id('submit').first.click()
                else: time.sleep(5)
            StrResult = browser.find_by_xpath('//div[@id="rsDiv"]').text
        except Exception as error : 
            print('GetPCCW : %s\n%s'%(Ip,error))
            DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
            receiver_email = 'thinhlv@vng.com.vn'
            subject = '[Ping monitor tool] Send error message'
            message = '[Function : Main][Time collect: %s][Error : %s]'%(DateTimeCurrent,error)
            SendEmailText(receiver_email,message,subject)
            time.sleep(60)

        if Ip in DctResult: DctResult[Ip]['PCCW'] = StrResult
        else: DctResult[Ip] = {'PCCW':StrResult}

def SendEmailAttachFile(receiver_email,subject,message,NameFileResult):
    dir_path = './/DataInfo'
    files = [NameFileResult]
    sender_email = 'tool.acl.thinhlv@gmail.com'
    password = 'Myt00l@cl'
    msg = MIMEMultipart()
    msg['To'] = receiver_email
    msg['From'] = sender_email
    msg['Subject'] = subject
    body = MIMEText(message, 'html', 'utf-8')  
    msg.attach(body)  # add message body (text or html)

    for f in files:  # add files to the message
        file_path = os.path.join(dir_path, f)
        attachment = MIMEApplication(open(file_path, "rb").read(), _subtype="txt")
        attachment.add_header('Content-Disposition','attachment', filename=f)
        msg.attach(attachment)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, msg.as_string()
    )

def SendEmailText (receiver_email,messages, subject):
    sender_email = "tool.acl.thinhlv@gmail.com"
    password = 'Myt00l@cl'
    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = receiver_email

    # Turn these into plain/html MIMEText objects
    part1 = MIMEText(messages, "plain")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part1)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
    )

if __name__ == "__main__":
  
    while True:
        Stop = True
        StrTime = input('Nhập vào list phút, mỗi phút cách nhau dấu phẩy or dấu cách : ').strip()
        if ',' in StrTime: LstTmp = StrTime.split(',')
        else: LstTmp = StrTime.split(' ')
        for Time in LstTmp: 
            if Time.isdigit() == False: Stop = False
        if Stop : break
        print('\n')
    LstTime = []
    for Time in LstTmp: LstTime.append(int(Time))
    LstTimeCheck = []
    while True: 
        TimeCurrent = datetime.datetime.now()
        DateTimeCurrent = str(datetime.datetime.now().strftime('%d:%m:%Y %H:%M:%S'))
        StrTimeCheck = '%s_%s_%s_%s'%(str(TimeCurrent.year),str(TimeCurrent.month),str(TimeCurrent.day),str(TimeCurrent.hour))
        if TimeCurrent.minute in LstTime and StrTimeCheck not in LstTimeCheck:
            try:
                StrTimeTmp = str(datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S'))
                NameFileResult = './/DataInfo/Report_%s.xlsx'%StrTimeTmp
                Wb = openpyxl.load_workbook(r'./DataInfo/LstIP.xlsx')
                SheetName = Wb.sheetnames
                Ws = Wb[SheetName[0]]
                LstIP,threads,DctResult = [],[],{}

                for IndexRow in range(2,Ws.max_row+1): LstIP.append(Ws.cell(row = IndexRow,column = 1).value)

                browser1 = Browser('chrome')
                browser2 = Browser('chrome')
                browser3 = Browser('chrome')
                x = threading.Thread(target=GetLgHeNet, args=(LstIP,browser1,DctResult))
                threads.append(x)
                x.start()
                x = threading.Thread(target=GetCenturyLink, args=(LstIP,browser2,DctResult))
                threads.append(x)
                x.start()
                x = threading.Thread(target=GetPCCW, args=(LstIP,browser3,DctResult))
                threads.append(x)
                x.start()
                for index, thread in enumerate(threads):
                    thread.join()
                browser1.quit()
                browser2.quit()
                browser3.quit()

                for IndexRow in range(2,Ws.max_row+1): 
                    Ip = Ws.cell(row = IndexRow,column = 1).value
                    LgHeNet, CenturyLink, PCCWglobal = ' ', ' ', ' '
                    if Ip in DctResult:
                        DctTmp = DctResult[Ip]
                        if 'LgHeNet' in DctTmp: LgHeNet = DctTmp['LgHeNet']
                        if 'CenturyLink' in DctTmp: CenturyLink = DctTmp['CenturyLink']
                        if 'PCCW' in DctTmp: PCCWglobal = DctTmp['PCCW']
                    
                    Ws.cell(row = IndexRow,column = 2).value = LgHeNet
                    Ws.cell(row = IndexRow,column = 3).value = CenturyLink
                    Ws.cell(row = IndexRow,column = 4).value = PCCWglobal
                    Ws.cell(row = IndexRow,column = 5).value = ' '

                Wb.save(NameFileResult)
                LstTimeCheck.append(StrTimeCheck)
                message = '<!DOCTYPE html><html><body><p>Time collect : %s<br><br>PING MONITOR TOOL</p></body></html>'%DateTimeCurrent
                receiver_email = 'thinhlv@vng.com.vn'
                subject = '[Ping monitor tool] Send file report'
                SendEmailAttachFile(receiver_email,subject,message,'Report_%s.xlsx'%StrTimeTmp)
            except Exception as error : 
                print('Main : %s'%error)
                receiver_email = 'thinhlv@vng.com.vn'
                subject = '[Ping monitor tool] Send error message'
                message = '[Function : Main][Time collect: %s][Error : %s]'%(DateTimeCurrent,error)
                SendEmailText(receiver_email,message,subject)
        else:
            print('List time collect mỗi ngày : %s'%(','.join(LstTmp)))
            print('Time current : %s'%DateTimeCurrent)
            time.sleep(30)